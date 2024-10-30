import argparse
import re
import yaml
import openpyxl
import xlsxwriter
from bs4 import BeautifulSoup, Tag
from copy import deepcopy
import inflect


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="""
        USDM Comformance Rules Test Data Template Generator.
        Reads the Enterprise Architect XMI export file containing definition
        of the USDM UML model (specified in the -x option), the USDM
        Controlled Terminology Excel file (specified in the -c option), and the
        USDM API specification in YAML format (specified in the -a option) and
        creates an Excel template file (optionally specified in the -o option)
        for test data entry."""
    )
    parser.add_argument(
        "-x",
        "--xmi_file",
        help="USDM XMI export file "
        + "(e.g.,<DDFversion>/Deliverables/UML/USDM_UML.xmi)",
        required=True,
    )
    parser.add_argument(
        "-c",
        "--ct_file",
        help="USDM CT Excel file "
        + "(e.g., <DDF version>/Deliverables/CT/USDM_CT.xlsx)",
        required=True,
    )
    parser.add_argument(
        "-a",
        "--api_spec",
        help="USDM API specification YAML file. "
        + "(e.g., <DDF version>/Deliverables/API/USDM_API.yaml)",
        required=True,
    )
    parser.add_argument(
        "-o",
        "--output_file",
        help="[Optional] Specifies output file Excel file. "
        + "Default is ./USDM_<USDM version>_Test_Data_Template.xlsx"
        + "(e.g., USDM_2.6_Test_Data_Template.xlsx)",
        default="USDM_<USDM version>_Test_Data_Template.xlsx",
    )
    args = parser.parse_args()
    return args


def replace_deep(data, a, b):
    if isinstance(data, str):
        return data.replace(a, b)
    elif isinstance(data, dict):
        return {k: replace_deep(v, a, b) for k, v in data.items()}
    elif isinstance(data, list):
        return [replace_deep(v, a, b) for v in data]
    else:
        return data


def get_api_desc(prp: str, prpDict: dict) -> str:
    if "type" in prpDict and prpDict["type"] == "array":
        return name_to_desc(prp) + " [Any Exist]"
    elif "$ref" in prpDict or (
        "anyOf" in prpDict and any("$ref" in x for x in prpDict["anyOf"])
    ):
        return name_to_desc(prp) + " [Exists]"
    else:
        return name_to_desc(prp)


def get_api_type(prpDict: dict) -> str:
    if (
        ("type" in prpDict and prpDict["type"] == "array")
        or "$ref" in prpDict
        or ("anyOf" in prpDict and any("$ref" in x for x in prpDict["anyOf"]))
    ):
        return "Boolean"
    elif "const" in prpDict:
        return "String"
    elif "type" in prpDict:
        return prpDict["type"].title()
    elif "anyOf" in prpDict:
        return "".join(
            x["type"].title()
            for x in prpDict["anyOf"]
            if "type" in x and x["type"] != "'null'"
        )
    else:
        return "!!UNKNOWN API TYPE!!"


def get_properties(
    entName: str, cls: Tag, prps: list, prefix: list = None, lclsName: str = None
):
    if cls.generalization:
        gclsId = cls.generalization["general"]
        gcls = usdmxmi.find("packagedElement", attrs={"xmi:id": gclsId})
        get_properties(entName, gcls, prps, prefix, cls["name"])
    for prp in (
        x
        for x in cls.find_all("ownedAttribute", attrs={"xmi:type": "uml:Property"})
        if x.has_attr("name")
    ):
        attr = usdmxmi.find("attribute", attrs={"xmi:idref": str(prp["xmi:id"])})
        prps[0] += [".".join((prefix[0], prp["name"])) if prefix else prp["name"]]
        prps[1] += [
            (
                " / ".join(
                    (prefix[1], get_description(lclsName or entName, prp["name"]))
                )
                if prefix
                else get_description(lclsName or entName, prp["name"])
            )
            + ""
            if attr.bounds["upper"] == "1"
            else " [Any Exist]"
        ]
        prps[2] += [
            attr.properties["type"] if attr.bounds["upper"] == "1" else "Boolean"
        ]
        prps[3] += [
            "{}.{}{}".format(prefix[2], prp["name"], get_card(attr.bounds))
            if prefix
            else get_card(attr.bounds)
        ]
    if not prefix and cls["name"] == entName:
        for prp in (
            x
            for x in apidict[entName].keys()
            if x != "id"
            and x not in entdict[entName]["Properties"]
            and not any(
                k
                for k, v in entdict[entName]["Properties"].items()
                if v["apiattr"] == x
            )
        ):
            prps[0] += [prp]
            prps[1] += [get_api_desc(prp, apidict[entName][prp])]
            prps[2] += [get_api_type(apidict[entName][prp])]
            prps[3] += ["[1]" if apidict[entName][prp]["required"] is True else "[0]"]
            print(
                (
                    "API-only attribute "
                    + f"'{entName}.{prp}' added from {args.api_spec}"
                )
            )

    for lnk in (
        x.find_parent("connector")
        for x in usdmxmi.find_all("source", attrs={"xmi:idref": cls["xmi:id"]})
        if x.find_parent("connector").has_attr("name")
    ):
        if lnk["name"] in entdict[lclsName if prefix else entName]["Properties"]:
            lcls = usdmxmi.find(
                "packagedElement", attrs={"xmi:id": lnk.target["xmi:idref"]}
            )
            if prefix:
                lnkName = ".".join((prefix[0], lnk["name"]))
                lnkDesc = " / ".join(
                    (
                        prefix[1],
                        get_description(lclsName or entName, lnk["name"]),
                    )
                )
                lnkCard = ">".join(
                    (
                        prefix[2],
                        "{}[{}]".format(
                            lcls["name"],
                            lnk.target.type["multiplicity"],
                        ),
                    )
                )
            else:
                lnkName = lnk["name"]
                lnkDesc = get_description(lclsName or entName, lnk["name"])
                lnkCard = "{}[{}]".format(lcls["name"], lnk.target.type["multiplicity"])
            apiattr = (
                entdict[lclsName if prefix else entName]["Properties"][lnk["name"]][
                    "apiattr"
                ]
                if "apiattr"
                in entdict[lclsName if prefix else entName]["Properties"][lnk["name"]]
                else None
            )
            single_lnk = lnk.target.type["multiplicity"].endswith("1")
            if apiattr is None or apiattr == lnk["name"]:
                isCircular: bool = False
                if prefix and lnk["name"] in prefix[0].split("."):
                    isCircular = True
                    print(
                        f"Circular relationship found: {lnk['name']} found in "
                        + prefix[0]
                    )
                else:
                    prps[0] += [lnkName]
                    prps[1] += [
                        lnkDesc + (" [Exists]" if single_lnk else " [Any Exists]")
                    ]
                    prps[2] += ["Boolean"]
                    prps[3] += [lnkCard]
                if single_lnk and not isCircular:
                    get_properties(
                        entName,
                        lcls,
                        prps,
                        [
                            lnkName,
                            lnkDesc,
                            lnkCard,
                        ],
                        lcls["name"],
                    )
            else:
                if (".".join((prefix[0], apiattr)) if prefix else apiattr) in prps[0]:
                    prps[3][
                        prps[0].index(
                            ".".join((prefix[0], apiattr)) if prefix else apiattr
                        )
                    ] += " / {}".format(
                        ">".join(
                            (
                                prefix[2],
                                "{}[{}].id[1]".format(
                                    lnk.target.model["name"],
                                    lnk.target.type["multiplicity"],
                                ),
                            )
                        )
                        if prefix
                        else "{}[{}].id[1]".format(
                            lnk.target.model["name"], lnk.target.type["multiplicity"]
                        )
                    )
                else:
                    prps[0] += [".".join((prefix[0], apiattr)) if prefix else apiattr]
                    prps[1] += [
                        "{} [{}]".format(
                            " / ".join(
                                (
                                    prefix[1],
                                    get_description(lclsName or entName, lnk["name"]),
                                )
                            )
                            if prefix
                            else get_description(lclsName or entName, lnk["name"]),
                            "Identifier" if single_lnk else "Identifiers][Any Exist",
                        )
                    ]
                    prps[2] += ["String" if single_lnk else "Boolean"]
                    prps[3] += [
                        ">".join(
                            (
                                prefix[2],
                                "{}[{}].id[1]".format(
                                    lnk.target.model["name"],
                                    lnk.target.type["multiplicity"],
                                ),
                            )
                        )
                        if prefix
                        else "{}[{}].id[1]".format(
                            lnk.target.model["name"], lnk.target.type["multiplicity"]
                        )
                    ]
        elif not prefix:
            print(
                f"Relationship '{entName}.{lnk['name']}' defined in {args.xmi_file} "
                + f"does not have a matching entry in {args.ct_file}"
            )


def get_description(entName: str, prpName: str = None) -> str:
    if (
        prpName in entdict[entName]["Properties"]
        and entdict[entName]["Properties"][prpName]["Preferred Name"]
    ):
        return entdict[entName]["Properties"][prpName]["Preferred Name"]
    else:
        if not (prpName in entdict[entName]["Properties"] or prpName == "id"):
            print(
                f"No entry found in {args.ct_file} for USDM "
                + f"attribute '{entName}.{prpName}'"
            )
        return "({} {})".format(name_to_desc(entName), name_to_desc(prpName))


def name_to_desc(name: str) -> str:
    return re.sub("([A-Z]+)", r" \1", name).strip().title()


def get_card(bounds: Tag) -> str:
    if bounds["lower"] == bounds["upper"]:
        return "[{}]".format(bounds["lower"])
    else:
        return "[{}..{}]".format(bounds["lower"], bounds["upper"])


def get_apiattr(entName: str, elname: str, elrole: str) -> str:
    if elname in apidict[entName]:
        return elname
    else:
        if elrole == "Attribute":
            print(f"No entry found in {args.api_spec} for '{entName}.{elname}'")
            return None
        else:
            elprts = re.findall(r"([A-Z]?[a-z]+)", elname.strip())
            if (
                inflect.singular_noun(elprts[-1]) is False
                or inflect.singular_noun(elprts[-1]) == elprts[-1]
            ):
                if elname + "Id" in apidict[entName]:
                    return elname + "Id"
                elif (
                    inflect.singular_noun(elprts[-1]) == elprts[-1]
                    and elname + "Ids" in apidict[entName]
                ):
                    return elname + "Ids"
                else:
                    print(
                        f"No entry found in {args.api_spec} for '{entName}."
                        + f"{elname}' or '{entName}.{elname}Id'"
                    )
            else:
                if len(elprts) == 1:
                    if inflect.singular_noun(elname) + "Ids" in apidict[entName]:
                        return inflect.singular_noun(elname) + "Ids"
                    elif elname + "Ids" in apidict[entName]:
                        print(
                            f"Using '{entName}.{elname}Ids' instead of "
                            + f"'{entName}.{inflect.singular_noun(elname)}Ids' "
                            + "for API attribute"
                        )
                        return elname + "Ids"
                    else:
                        print(
                            f"No entry found in {args.api_spec} for "
                            + f"'{entName}.{elname}', '{entName}."
                            + f"{inflect.singular_noun(elname)}Ids' or "
                            + f"'{entName}.{elname}Ids'"
                        )
                        return None
                else:
                    if (
                        "".join(elprts[:-1]) + inflect.singular_noun(elprts[-1]) + "Ids"
                        in apidict[entName]
                    ):
                        return (
                            "".join(elprts[:-1])
                            + inflect.singular_noun(elprts[-1])
                            + "Ids"
                        )
                    elif elname + "Ids" in apidict[entName]:
                        print(
                            f"Using '{entName}.{elname}Ids' instead of "
                            + f"'{entName}."
                            + "".join(elprts[:-1])
                            + inflect.singular_noun(elprts[-1])
                            + "Ids' for API attribute"
                        )
                        return elname + "Ids"
                    else:
                        print(
                            f"No entry found in {args.api_spec} for '{entName}."
                            + f"{elname}', '{entName}."
                            + "".join(elprts[:-1])
                            + inflect.singular_noun(elprts[-1])
                            + f"Ids' or '{entName}.{elname}Ids'"
                        )
                        return None


args = parse_arguments()

with open(args.xmi_file) as f:
    xmidata = f.read()

usdmxmi = BeautifulSoup(xmidata, "lxml-xml")

with open(args.api_spec, "r") as f:
    apispec = yaml.safe_load(f)

apidict = {}

for k, v in apispec["components"]["schemas"].items():
    if "-" in k:
        if k.endswith("-Input"):
            cname = "".join(k.split("-")[:1])
            if cname + "-Output" in apispec["components"]["schemas"]:
                vo = apispec["components"]["schemas"][cname + "-Output"]
                if v != replace_deep(
                    vo,
                    "Output",
                    "Input",
                ):
                    print(f"API Input/Output definitions do not match for {cname}")
                    print(f"{cname}-Input : {v}")
                    print(f"{cname}-Output: {vo}")
            else:
                print(f"No corresponding API Output definition for {k}")
    else:
        cname = k

    apidict[cname] = deepcopy(v["properties"])
    for apiattn, apiattv in apidict[cname].items():
        apiattv["required"] = "required" in v and apiattn in v["required"]

inflect = inflect.engine()
inflect.defnoun("previous", "previous")
inflect.defnoun("context", "context")
inflect.defnoun("to", "to")
inflect.defnoun("of", "of")

entdict = {}

ctwb = openpyxl.load_workbook(filename=args.ct_file, data_only=True)

ctws = ctwb["DDF Entities&Attributes"]

ctcolmap = {ctcol.value: ctcol.column - 1 for ctcol in tuple(ctws.rows)[0]}

for ctrow in ctws.iter_rows(min_row=2):
    entName = ctrow[ctcolmap["Entity Name"]].value
    elrole = ctrow[ctcolmap["Role"]].value
    elname = ctrow[ctcolmap["Logical Data Model Name"]].value
    if elrole == "Entity":
        if elname != entName:
            print(
                f"Entity Name '{entName}' does not match Logical Data Model "
                + f"Name for Entity '{elname}'"
            )
        entdict[entName] = {
            "NCI C-code": ctrow[ctcolmap["NCI C-code"]].value,
            "Preferred Name": ctrow[ctcolmap["CT Item Preferred Name"]].value,
            "Definition": ctrow[ctcolmap["Definition"]].value,
            "Properties": {},
        }
    else:
        cref: str = None
        cref = re.search(
            r"^Y \((.+?)\)$", str(ctrow[ctcolmap["Has Value List"]].value).strip()
        )

        entdict[entName]["Properties"][elname] = {
            "name": elname,
            "Role": elrole,
            "NCI C-code": ctrow[ctcolmap["NCI C-code"]].value,
            "Preferred Name": ctrow[ctcolmap["CT Item Preferred Name"]].value,
            "Definition": ctrow[ctcolmap["CT Item Preferred Name"]].value,
            "CodelistRef": cref.group(1) if cref else None,
        }

for entName, entDef in entdict.items():
    cls = usdmxmi.find(
        "packagedElement", attrs={"xmi:type": "uml:Class", "name": entName}
    )
    if cls.generalization:
        gclsName = usdmxmi.find(
            "packagedElement",
            attrs={
                "xmi:type": "uml:Class",
                "xmi:id": cls.generalization["general"],
            },
        )["name"]
        if gclsName in entdict:
            for gprp in entdict[gclsName]["Properties"].keys():
                if gprp not in entDef["Properties"]:
                    entDef["Properties"][gprp] = deepcopy(
                        entdict[gclsName]["Properties"][gprp]
                    )
                    print(
                        f"Using general '{gclsName}.{gprp}' "
                        + entDef["Properties"][gprp]["Role"].lower()
                        + f" in '{entName}' specialization"
                    )
    if entName in apidict:
        for prpName, prpDef in entDef["Properties"].items():
            prpDef["apiattr"] = get_apiattr(entName, prpName, prpDef["Role"])

for abscls in (
    x["name"]
    for x in usdmxmi.find_all("element", attrs={"xmi:type": "uml:Class"})
    if x.properties["isAbstract"] == "true" or x["name"] not in apidict
):
    if abscls in entdict:
        entdict.pop(abscls)
        print(f"Excluding abstract class '{abscls}'")
    else:
        print(f"Abstract class {abscls} not found in {args.ct_file}")

for clsName, clsDef in apidict.items():
    if clsName not in entdict:
        print(
            f"No entry found in {args.ct_file} for API class '{clsName}' from "
            + args.api_spec
        )
    else:
        for prp in (
            x
            for x in clsDef.keys()
            if x != "id"
            and x not in entdict[clsName]["Properties"]
            and not any(
                k
                for k, v in entdict[clsName]["Properties"].items()
                if v["apiattr"] == x
            )
        ):
            print(
                (
                    f"No corresponding entry found in {args.ct_file} for API attribute "
                    + f"'{clsName}.{prp}' from {args.api_spec}"
                )
            )

usdmver = (
    usdmxmi.find("properties", {"name": "USDM", "type": "Logical"})
    .find_parent("diagram")
    .project["version"]
)

workbook = xlsxwriter.Workbook(args.output_file.replace("<USDM version>", usdmver))
workbook.set_custom_property("USDM Version", str(usdmver))

header = workbook.add_format()
header.set_bold()
header.set_align("top")
header.set_text_wrap()

sub_header = workbook.add_format()
sub_header.set_italic()
sub_header.set_bg_color("#FFFFCC")
sub_header.set_text_wrap()
sub_header.set_align("top")

normal = workbook.add_format()
normal.set_align("top")
normal.set_num_format("@")

dsws = workbook.add_worksheet("Datasets")
dsprps = ["Filename", "Dataset Name", "Label"]
dsws.set_column(0, len(dsprps), 30)
dsws.write_row(0, 0, dsprps, header)

clsn = 0

for entName in entdict.keys():
    cls = usdmxmi.find(
        "packagedElement", attrs={"xmi:type": "uml:Class", "name": entName}
    )
    if cls:
        clsn += 1
        clsSheet = entName + ".xpt" if len(entName) <= 27 else entName[:27] + ".xpt"
        dsws.write_url(clsn, 0, f"internal:'{clsSheet}'!A1", string=clsSheet)
        dsws.write_row(clsn, 1, [entName, entdict[entName]["Preferred Name"]])
        ws = workbook.add_worksheet(clsSheet)
        prps = [[], [], [], []]
        get_properties(entName, cls, prps)
        for prpv in (
            v
            for k, v in entdict[entName]["Properties"].items()
            if not (
                cls.find(
                    "ownedAttribute", attrs={"xmi:type": "uml:Property", "name": k}
                )
                or any(
                    x
                    for x in usdmxmi.find_all(
                        "source", attrs={"xmi:idref": cls["xmi:id"]}
                    )
                    if x.find_parent("connector").has_attr("name")
                    and x.find_parent("connector")["name"] == k
                )
                or (
                    cls.generalization
                    and (
                        usdmxmi.find(
                            "packagedElement",
                            attrs={
                                "xmi:type": "uml:Class",
                                "xmi:id": cls.generalization["general"],
                            },
                        ).find(
                            "ownedAttribute",
                            attrs={"xmi:type": "uml:Property", "name": k},
                        )
                        or any(
                            x
                            for x in usdmxmi.find_all(
                                "source",
                                attrs={"xmi:idref": cls.generalization["general"]},
                            )
                            if x.find_parent("connector").has_attr("name")
                            and x.find_parent("connector")["name"] == k
                        )
                    )
                )
            )
        ):
            print(
                f"{prpv['Role']} '{entName}.{prpv['name']}' defined in {args.ct_file} "
                + "does not have a matching attribute or relationship in "
                + args.xmi_file
            )
        prps[0] = ["parent_entity", "parent_id", "parent_rel", "rel_type"] + prps[0]
        prps[1] = [
            "Parent Entity Name",
            "Parent Entity Id",
            "Name of Relationship from Parent Entity",
            "Type of Relationship",
        ] + prps[1]
        prps[2] = ["String", "String", "String", "String"] + prps[2]
        prps[3] = ["[1]", "[1]", "[1]", "[1]"] + prps[3]
        ws.set_column(0, len(prps[0]), 25)
        ws.write_row(0, 0, prps[0], header)
        ws.write_row(1, 0, prps[1], sub_header)
        ws.write_row(2, 0, prps[2], sub_header)
        ws.write_row(3, 0, prps[3], sub_header)
        # Add a blank row with defined format to prevent auto-copying of format
        # from row above.
        ws.write_row(4, 0, [None] * len(prps[0]), normal)
    else:
        print(
            f"Entity '{entName}' defined in {args.ct_file} does not have a "
            + f"matching class in {args.xmi_file}"
        )

for cls in (
    x
    for x in usdmxmi.find_all("packagedElement", attrs={"xmi:type": "uml:Class"})
    if x["name"] not in entdict
    and usdmxmi.find("element", attrs={"xmi:idref": x["xmi:id"]}).properties[
        "isAbstract"
    ]
    == "false"
    and x["name"] in apidict
):
    print(
        f"USDM class '{cls['name']}' defined in {args.xmi_file} does not "
        + f"have a matching Entity in {args.ct_file}"
    )

for pdtype in ["String", "Float", "Boolean", "Null"]:
    clsn += 1
    dsname = pdtype.lower()
    dsws.write_url(clsn, 0, f"internal:'{dsname}.xpt'!A1", string=f"{dsname}.xpt")
    dsws.write_row(clsn, 1, [dsname, f"{pdtype} Values"])
    ws = workbook.add_worksheet(f"{dsname}.xpt")
    ws.set_column(0, 4, 25)
    ws.write_row(
        0, 0, ["parent_entity", "parent_id", "parent_rel", "rel_type", "value"], header
    )
    ws.write_row(
        1,
        0,
        [
            "Parent Entity Name",
            "Parent Entity Id",
            "Name of Relationship from Parent Entity",
            "Type of Relationship",
            "Value",
        ],
        sub_header,
    )
    ws.write_row(2, 0, ["String"] * 5, sub_header)
    ws.write_row(3, 0, ["[1]"] * 4 + ["[0]" if pdtype == "Null" else "[1]"], sub_header)
    # Add a blank row with defined format to prevent auto-copying of format
    # from row above.
    ws.write_row(4, 0, [None] * 5, normal)

workbook.close()
